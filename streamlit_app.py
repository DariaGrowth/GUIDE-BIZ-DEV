# =============================================================================
# INGOOD BY OLGA — Sulfodyne Prospect Tracker
# Version 1.0 | Business Development Tool
# =============================================================================

import streamlit as st
import pandas as pd
import numpy as np
from supabase import create_client
import google.generativeai as genai
from datetime import datetime
import urllib.parse
from io import BytesIO
import time

# =============================================================================
# PAGE CONFIG
# =============================================================================

FAVICON_SVG = """<svg width="32" height="32" viewBox="0 0 56 56" fill="none" xmlns="http://www.w3.org/2000/svg"><rect width="56" height="56" rx="12" fill="#1E3F35"/><path d="M28 12c-3 7-9 11-16 11 7 3 13 9 16 16 3-7 9-13 16-16-7-3-13-6-16-11z" fill="white"/></svg>"""

st.set_page_config(
    page_title="Sulfodyne Tracker · Ingood",
    page_icon="🌿",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={"Get help": None, "Report a bug": None, "About": None},
)

# =============================================================================
# DESIGN SYSTEM
# =============================================================================

CSS = """
<style>
    @import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;500;600;700;800&family=DM+Mono:wght@400;500&family=DM+Sans:wght@400;500;600&display=swap');

    :root {
        --forest:    #1E3F35;
        --forest-2:  #2A5548;
        --moss:      #10B981;
        --lime:      #84CC16;
        --cream:     #FAFAF7;
        --white:     #FFFFFF;
        --border:    #E8EAE6;
        --border-2:  #D1D5CC;
        --text:      #1A1F1C;
        --text-2:    #4A5450;
        --text-3:    #8A9490;
        --high:      #DC2626;
        --mid:       #F59E0B;
        --low:       #6B7280;
        --high-bg:   #FEF2F2;
        --mid-bg:    #FFFBEB;
        --low-bg:    #F9FAFB;
    }

    /* ── BASE ── */
    .stApp {
        background: var(--cream) !important;
        font-family: 'DM Sans', sans-serif !important;
    }
    * { font-family: 'DM Sans', sans-serif !important; box-sizing: border-box !important; }
    h1, h2, h3 { font-family: 'Syne', sans-serif !important; }
    #MainMenu, footer, header { visibility: hidden; }
    .stDeployButton, div[data-testid="stToolbar"] { display: none !important; }
    div[data-testid="stVerticalBlockBorderWrapper"] { border: none !important; background: transparent !important; padding: 0 !important; }
    [data-testid="stVerticalBlock"] { gap: 0 !important; }
    button:focus, button:focus-visible { outline: none !important; box-shadow: none !important; }

    /* ── SIDEBAR ── */
    section[data-testid="stSidebar"] {
        background: var(--forest) !important;
        border-right: none !important;
        min-width: 220px !important;
        max-width: 220px !important;
    }
    section[data-testid="stSidebar"] > div { padding: 28px 16px !important; }
    section[data-testid="stSidebar"] * { color: rgba(255,255,255,0.85) !important; }
    section[data-testid="stSidebar"] div[data-testid="stVerticalBlockBorderWrapper"],
    section[data-testid="stSidebar"] .stButton {
        background: transparent !important; border: none !important; box-shadow: none !important;
    }

    /* Sidebar primary button */
    section[data-testid="stSidebar"] .stButton > button[kind="primary"] {
        width: 100% !important;
        background: rgba(255,255,255,0.12) !important;
        color: white !important;
        border: 1px solid rgba(255,255,255,0.2) !important;
        border-radius: 10px !important;
        padding: 13px 16px !important;
        font-family: 'Syne', sans-serif !important;
        font-weight: 700 !important;
        font-size: 13px !important;
        letter-spacing: 0.3px !important;
        transition: all 0.15s ease !important;
    }
    section[data-testid="stSidebar"] .stButton > button[kind="primary"]:hover {
        background: rgba(255,255,255,0.2) !important;
        border-color: rgba(255,255,255,0.35) !important;
    }

    /* Sidebar secondary nav buttons */
    section[data-testid="stSidebar"] .stButton > button[kind="secondary"] {
        width: 100% !important;
        background: transparent !important;
        color: rgba(255,255,255,0.65) !important;
        border: none !important;
        border-radius: 8px !important;
        padding: 11px 14px !important;
        font-size: 13.5px !important;
        font-weight: 500 !important;
        text-align: left !important;
        justify-content: flex-start !important;
        transition: all 0.15s ease !important;
        margin: 1px 0 !important;
    }
    section[data-testid="stSidebar"] .stButton > button[kind="secondary"]:hover {
        background: rgba(255,255,255,0.1) !important;
        color: white !important;
    }

    /* ── MAIN CONTENT ── */
    .block-container { padding: 32px 40px 40px 40px !important; max-width: 1400px !important; }

    /* ── PAGE HEADER ── */
    .page-header {
        margin-bottom: 28px;
    }
    .page-title {
        font-family: 'Syne', sans-serif !important;
        font-size: 26px !important;
        font-weight: 800 !important;
        color: var(--text) !important;
        margin: 0 0 4px 0 !important;
        letter-spacing: -0.3px !important;
    }
    .page-subtitle {
        font-size: 14px !important;
        color: var(--text-3) !important;
        margin: 0 !important;
    }

    /* ── PROSPECT TABLE ── */
    .table-wrap {
        background: var(--white);
        border: 1px solid var(--border);
        border-radius: 14px;
        overflow: hidden;
    }
    .table-top {
        display: flex;
        justify-content: space-between;
        align-items: center;
        padding: 18px 24px;
        border-bottom: 1px solid var(--border);
    }
    .table-count {
        font-size: 13px;
        color: var(--text-3);
        font-weight: 500;
    }

    /* Table header row */
    .th-row {
        display: grid;
        grid-template-columns: 2fr 0.7fr 1fr 1.4fr 1.2fr 1.3fr 1fr 60px;
        padding: 10px 24px;
        background: #F6F7F5;
        border-bottom: 1px solid var(--border);
        gap: 12px;
    }
    .th-cell {
        font-size: 10.5px !important;
        font-weight: 700 !important;
        color: var(--text-3) !important;
        text-transform: uppercase !important;
        letter-spacing: 0.6px !important;
    }

    /* Table data rows */
    .td-row {
        display: grid;
        grid-template-columns: 2fr 0.7fr 1fr 1.4fr 1.2fr 1.3fr 1fr 60px;
        padding: 14px 24px;
        border-bottom: 1px solid #F0F1EF;
        align-items: center;
        gap: 12px;
        transition: background 0.12s ease;
    }
    .td-row:hover { background: #FAFBF9; }
    .td-row:last-child { border-bottom: none; }

    .td-company {
        font-weight: 700 !important;
        font-size: 14px !important;
        color: var(--text) !important;
        text-transform: uppercase !important;
        letter-spacing: 0.2px !important;
    }
    .td-text {
        font-size: 13px !important;
        color: var(--text-2) !important;
    }
    .td-link {
        font-size: 12px !important;
        color: var(--forest) !important;
        text-decoration: none !important;
        font-weight: 500 !important;
    }
    .td-link:hover { text-decoration: underline !important; }
    .td-mono {
        font-family: 'DM Mono', monospace !important;
        font-size: 12px !important;
        color: var(--text-2) !important;
    }
    .td-analyse {
        font-size: 12px !important;
        color: var(--text-2) !important;
        line-height: 1.4 !important;
        display: -webkit-box !important;
        -webkit-line-clamp: 2 !important;
        -webkit-box-orient: vertical !important;
        overflow: hidden !important;
    }

    /* Priority badges */
    .badge-high { display: inline-flex; align-items: center; gap: 5px; background: var(--high-bg); color: var(--high); font-size: 11px; font-weight: 700; padding: 4px 10px; border-radius: 20px; letter-spacing: 0.3px; }
    .badge-mid  { display: inline-flex; align-items: center; gap: 5px; background: var(--mid-bg);  color: #92400E; font-size: 11px; font-weight: 700; padding: 4px 10px; border-radius: 20px; letter-spacing: 0.3px; }
    .badge-low  { display: inline-flex; align-items: center; gap: 5px; background: var(--low-bg);  color: var(--low); font-size: 11px; font-weight: 700; padding: 4px 10px; border-radius: 20px; letter-spacing: 0.3px; }

    /* ── MODAL / DIALOG ── */
    div[data-testid="stDialog"] > div > div {
        background: var(--white) !important;
        border-radius: 16px !important;
        box-shadow: 0 25px 60px -12px rgba(0,0,0,0.22) !important;
        max-width: 780px !important;
        width: 92vw !important;
        padding: 32px !important;
    }
    div[data-testid="stDialog"] > div:first-child {
        background: rgba(0,0,0,0.35) !important;
        backdrop-filter: blur(5px) !important;
    }

    /* ── FORM ── */
    .form-label {
        font-family: 'Syne', sans-serif !important;
        font-size: 11px !important;
        font-weight: 700 !important;
        color: var(--text-2) !important;
        text-transform: uppercase !important;
        letter-spacing: 0.6px !important;
        margin: 0 0 6px 0 !important;
        display: block !important;
    }
    .stTextInput input, .stTextArea textarea, .stSelectbox > div > div {
        border: 1.5px solid var(--border) !important;
        border-radius: 8px !important;
        font-size: 14px !important;
        padding: 11px 13px !important;
        background: var(--white) !important;
        transition: border-color 0.15s ease !important;
    }
    .stTextInput input:focus, .stTextArea textarea:focus {
        border-color: var(--forest) !important;
        box-shadow: 0 0 0 3px rgba(30,63,53,0.08) !important;
    }

    /* ── METRICS ── */
    [data-testid="stMetric"] {
        background: var(--white) !important;
        border: 1px solid var(--border) !important;
        border-radius: 12px !important;
        padding: 20px 24px !important;
    }
    [data-testid="stMetricValue"] { font-family: 'Syne', sans-serif !important; font-size: 30px !important; font-weight: 800 !important; color: var(--text) !important; }
    [data-testid="stMetricLabel"] { font-size: 11px !important; font-weight: 700 !important; color: var(--text-3) !important; text-transform: uppercase !important; letter-spacing: 0.5px !important; }

    /* ── BUTTONS (main content) ── */
    .stButton > button[kind="primary"] {
        background: var(--forest) !important;
        color: white !important;
        border: none !important;
        border-radius: 8px !important;
        font-weight: 600 !important;
        font-size: 14px !important;
        padding: 11px 22px !important;
        transition: background 0.15s ease !important;
    }
    .stButton > button[kind="primary"]:hover {
        background: var(--forest-2) !important;
    }
    .stButton > button[kind="secondary"] {
        background: var(--white) !important;
        color: var(--text-2) !important;
        border: 1.5px solid var(--border) !important;
        border-radius: 8px !important;
        font-weight: 500 !important;
        font-size: 14px !important;
        padding: 11px 22px !important;
        transition: all 0.15s ease !important;
    }
    .stButton > button[kind="secondary"]:hover {
        border-color: var(--forest) !important;
        color: var(--forest) !important;
        background: #F0F9F5 !important;
    }

    /* ── SEARCH / FILTER ── */
    .stTextInput input[placeholder*="Rechercher"],
    .stTextInput input[placeholder*="Search"] {
        background: #F6F7F5 !important;
        border-color: transparent !important;
    }

    /* ── NEWS CARDS ── */
    .news-card {
        background: var(--white);
        border: 1px solid var(--border);
        border-radius: 12px;
        padding: 20px 22px;
        margin-bottom: 12px;
        transition: box-shadow 0.15s ease;
    }
    .news-card:hover { box-shadow: 0 4px 16px rgba(0,0,0,0.07); }
    .news-tag {
        display: inline-block;
        background: #ECFDF5;
        color: #065F46;
        font-size: 10px;
        font-weight: 700;
        text-transform: uppercase;
        letter-spacing: 0.5px;
        padding: 3px 9px;
        border-radius: 20px;
        margin-bottom: 10px;
    }
    .news-title {
        font-family: 'Syne', sans-serif !important;
        font-size: 15px !important;
        font-weight: 700 !important;
        color: var(--text) !important;
        margin: 0 0 8px 0 !important;
        line-height: 1.4 !important;
    }
    .news-body {
        font-size: 13px !important;
        color: var(--text-2) !important;
        line-height: 1.6 !important;
        margin: 0 !important;
    }
    .news-meta {
        font-size: 11px !important;
        color: var(--text-3) !important;
        margin-top: 10px !important;
        font-family: 'DM Mono', monospace !important;
    }

    /* ── EMPTY STATE ── */
    .empty-state {
        text-align: center;
        padding: 60px 20px;
        color: var(--text-3);
    }
    .empty-icon { font-size: 40px; margin-bottom: 12px; }
    .empty-title { font-family: 'Syne', sans-serif; font-size: 16px; font-weight: 700; color: var(--text-2); margin-bottom: 6px; }
    .empty-sub { font-size: 13px; color: var(--text-3); }

    /* ── DIVIDERS ── */
    hr { border: none !important; border-top: 1px solid var(--border) !important; margin: 24px 0 !important; }

    /* ── TABS ── */
    .stTabs [data-baseweb="tab-list"] { gap: 0 !important; border-bottom: 1px solid var(--border) !important; background: transparent !important; }
    .stTabs [data-baseweb="tab"] { font-size: 13px !important; font-weight: 500 !important; color: var(--text-3) !important; padding: 12px 20px !important; border-bottom: 2px solid transparent !important; background: transparent !important; }
    .stTabs [data-baseweb="tab"][aria-selected="true"] { color: var(--forest) !important; border-bottom-color: var(--forest) !important; font-weight: 700 !important; }
    .stTabs [data-baseweb="tab"]:hover { background: #F0F9F5 !important; color: var(--forest) !important; }
</style>
"""
st.markdown(CSS, unsafe_allow_html=True)

# =============================================================================
# CONNEXIONS
# =============================================================================

@st.cache_resource
def init_connections():
    try:
        url  = st.secrets["SUPABASE_URL"]
        key  = st.secrets["SUPABASE_KEY"]
        genai.configure(api_key=st.secrets["GOOGLE_API_KEY"])
        return create_client(url, key)
    except Exception as e:
        st.error(f"Connexion échouée : {e}")
        return None

def db():
    if "supabase" not in st.session_state:
        st.session_state.supabase = init_connections()
    return st.session_state.supabase

# =============================================================================
# HELPERS
# =============================================================================

PRIORITY_OPTS = ["High", "Middle", "Low"]

def priority_badge(p):
    if p == "High":   return '<span class="badge-high">● High</span>'
    if p == "Middle": return '<span class="badge-mid">● Middle</span>'
    return '<span class="badge-low">● Low</span>'

def safe_del(k):
    if k in st.session_state: del st.session_state[k]

def refresh():
    st.cache_data.clear()

# =============================================================================
# DATA LAYER
# =============================================================================

@st.cache_data(ttl=30)
def load_prospects():
    try:
        res = db().table("sulfodyne_prospects").select("*").order("created_at", desc=True).execute()
        return pd.DataFrame(res.data) if res.data else pd.DataFrame()
    except:
        return pd.DataFrame()

def upsert_prospect(data: dict, pid=None):
    data["updated_at"] = datetime.now().isoformat()
    if pid:
        db().table("sulfodyne_prospects").update(data).eq("id", pid).execute()
    else:
        data["created_at"] = datetime.now().isoformat()
        db().table("sulfodyne_prospects").insert(data).execute()
    refresh()

def delete_prospect(pid):
    db().table("sulfodyne_prospects").delete().eq("id", pid).execute()
    refresh()

# =============================================================================
# MODAL — ADD / EDIT
# =============================================================================

@st.dialog("Fiche Prospect", width="large")
def prospect_modal(pid=None, initial_data=None):
    d = initial_data or {}
    is_edit = pid is not None

    st.markdown(f"""
        <h2 style="font-family:'Syne',sans-serif;font-size:20px;font-weight:800;color:#1A1F1C;margin:0 0 4px;">
            {'✎  Modifier le prospect' if is_edit else '✦  Nouveau Prospect'}
        </h2>
        <p style="font-size:13px;color:#8A9490;margin:0 0 24px;">
            {'Mettez à jour les informations.' if is_edit else 'Renseignez les informations du prospect.'}
        </p>
    """, unsafe_allow_html=True)

    r1c1, r1c2 = st.columns([2, 1], gap="medium")
    with r1c1:
        st.markdown('<span class="form-label">SOCIÉTÉ *</span>', unsafe_allow_html=True)
        company = st.text_input("company", value=d.get("company",""), label_visibility="collapsed", placeholder="ex: Laboratoire XYZ")
    with r1c2:
        st.markdown('<span class="form-label">PRIORITÉ</span>', unsafe_allow_html=True)
        pri_idx = PRIORITY_OPTS.index(d.get("priority","High")) if d.get("priority") in PRIORITY_OPTS else 0
        priority = st.selectbox("priority", PRIORITY_OPTS, index=pri_idx, label_visibility="collapsed")

    st.markdown("<div style='height:16px;'></div>", unsafe_allow_html=True)

    r2c1, r2c2 = st.columns(2, gap="medium")
    with r2c1:
        st.markdown('<span class="form-label">PRODUIT / CATÉGORIE</span>', unsafe_allow_html=True)
        product = st.text_input("product", value=d.get("product",""), label_visibility="collapsed", placeholder="ex: Complément alimentaire, BAD...")
    with r2c2:
        st.markdown('<span class="form-label">DECISION MAKER</span>', unsafe_allow_html=True)
        dm = st.text_input("dm", value=d.get("decision_maker",""), label_visibility="collapsed", placeholder="Nom, Titre")

    st.markdown("<div style='height:16px;'></div>", unsafe_allow_html=True)

    r3c1, r3c2 = st.columns(2, gap="medium")
    with r3c1:
        st.markdown('<span class="form-label">EMAIL</span>', unsafe_allow_html=True)
        email = st.text_input("email", value=d.get("email",""), label_visibility="collapsed", placeholder="contact@company.com")
    with r3c2:
        st.markdown('<span class="form-label">WEBSITE</span>', unsafe_allow_html=True)
        website = st.text_input("website", value=d.get("website",""), label_visibility="collapsed", placeholder="https://...")

    st.markdown("<div style='height:16px;'></div>", unsafe_allow_html=True)

    st.markdown('<span class="form-label">ANALYSE / NOTES</span>', unsafe_allow_html=True)
    analyse = st.text_area("analyse", value=d.get("analyse",""), height=100, label_visibility="collapsed",
                           placeholder="Contexte, opportunité, pain points, historique des échanges...")

    st.markdown("<hr>", unsafe_allow_html=True)

    fc1, fc2, fc3 = st.columns([2, 1, 1])
    with fc2:
        if st.button("Annuler", key="modal_cancel", use_container_width=True, type="secondary"):
            st.rerun()
    with fc3:
        if st.button("✓ Enregistrer", key="modal_save", use_container_width=True, type="primary"):
            if not company.strip():
                st.error("Le nom de la société est obligatoire")
            else:
                payload = {
                    "company": company.strip().upper(),
                    "priority": priority,
                    "product": product,
                    "decision_maker": dm,
                    "email": email,
                    "website": website,
                    "analyse": analyse,
                }
                upsert_prospect(payload, pid)
                st.success("✓ Enregistré !")
                time.sleep(0.5)
                safe_del("edit_pid")
                st.rerun()

    # Delete zone (edit only)
    if is_edit:
        st.markdown("---")
        if st.session_state.get(f"confirm_del_{pid}"):
            st.warning("⚠️ Supprimer définitivement ce prospect ?")
            dc1, dc2 = st.columns(2)
            with dc1:
                if st.button("Confirmer la suppression", key="del_confirm", type="primary", use_container_width=True):
                    delete_prospect(pid)
                    safe_del(f"confirm_del_{pid}")
                    safe_del("edit_pid")
                    st.rerun()
            with dc2:
                if st.button("Annuler", key="del_cancel", use_container_width=True):
                    safe_del(f"confirm_del_{pid}")
                    st.rerun()
        else:
            if st.button("🗑  Supprimer ce prospect", key=f"del_init_{pid}", type="secondary"):
                st.session_state[f"confirm_del_{pid}"] = True
                st.rerun()

# =============================================================================
# SIDEBAR
# =============================================================================

def render_sidebar():
    with st.sidebar:
        st.markdown("""
            <div style="text-align:center; padding: 4px 0 24px;">
                <div style="font-family:'Syne',sans-serif; font-size:19px; font-weight:800; color:white; letter-spacing:-0.3px;">Ingood by Olga</div>
                <div style="font-size:11px; color:rgba(255,255,255,0.45); text-transform:uppercase; letter-spacing:1.5px; margin-top:4px;">Sulfodyne · BD</div>
            </div>
        """, unsafe_allow_html=True)

        if st.button("✦  Nouveau Prospect", key="btn_new", use_container_width=True, type="primary"):
            st.session_state["open_new"] = True
            st.rerun()

        st.markdown("<div style='height:20px;'></div>", unsafe_allow_html=True)

        if "nav_page" not in st.session_state:
            st.session_state.nav_page = "Prospects"

        pages = [
            ("Prospects",    "🧭"),
            ("Veille IA",    "🔬"),
            ("Import/Export","📊"),
        ]

        for label, icon in pages:
            is_active = st.session_state.nav_page == label
            btn_style = "primary" if is_active else "secondary"
            if st.button(f"{icon}  {label}", key=f"nav_{label}", use_container_width=True, type="secondary"):
                st.session_state.nav_page = label
                st.rerun()

        # Active state styling hack
        st.markdown(f"""
            <style>
                /* Highlight active nav */
                section[data-testid="stSidebar"] button[kind="secondary"] {{
                    color: rgba(255,255,255,0.65) !important;
                }}
            </style>
        """, unsafe_allow_html=True)

        st.markdown("<div style='flex:1; min-height:60px;'></div>", unsafe_allow_html=True)
        st.markdown("---")
        df = load_prospects()
        total = len(df)
        high  = len(df[df["priority"]=="High"]) if not df.empty and "priority" in df.columns else 0
        st.markdown(f"""
            <div style="font-size:12px; color:rgba(255,255,255,0.4); line-height:1.8;">
                <div>{total} prospect{'s' if total!=1 else ''}</div>
                <div style="color:#EF4444;">{high} High priority</div>
            </div>
        """, unsafe_allow_html=True)

    return st.session_state.nav_page

# =============================================================================
# PAGE: PROSPECTS
# =============================================================================

def page_prospects():
    df = load_prospects()

    # Header
    st.markdown("""
        <div class="page-header">
            <h1 class="page-title">Prospects Sulfodyne</h1>
            <p class="page-subtitle">Suivi Business Development · Ingood by Olga</p>
        </div>
    """, unsafe_allow_html=True)

    # KPIs
    if not df.empty:
        k1, k2, k3, k4 = st.columns(4)
        k1.metric("Total Prospects", len(df))
        k2.metric("Priorité High", len(df[df.get("priority","") == "High"]) if "priority" in df.columns else 0)
        k3.metric("Priorité Middle", len(df[df.get("priority","") == "Middle"]) if "priority" in df.columns else 0)
        k4.metric("Priorité Low", len(df[df.get("priority","") == "Low"]) if "priority" in df.columns else 0)
        st.markdown("<div style='height:24px;'></div>", unsafe_allow_html=True)

    # Filter bar
    fc1, fc2, fc3 = st.columns([3, 1.5, 1.5])
    with fc1:
        search = st.text_input("search", placeholder="🔍  Rechercher un prospect…", label_visibility="collapsed")
    with fc2:
        prio_filter = st.selectbox("pf", ["Toutes Priorités"] + PRIORITY_OPTS, label_visibility="collapsed")
    with fc3:
        if df.empty or "product" not in df.columns:
            products_filter = st.selectbox("prdf", ["Tous Produits"], label_visibility="collapsed")
        else:
            prod_list = ["Tous Produits"] + sorted(df["product"].dropna().unique().tolist())
            products_filter = st.selectbox("prdf", prod_list, label_visibility="collapsed")

    # Apply filters
    filtered = df.copy() if not df.empty else pd.DataFrame()
    if not filtered.empty:
        if search:
            mask = (
                filtered.get("company","").str.contains(search, case=False, na=False) |
                filtered.get("decision_maker","").str.contains(search, case=False, na=False) |
                filtered.get("analyse","").str.contains(search, case=False, na=False)
            )
            filtered = filtered[mask]
        if prio_filter != "Toutes Priorités":
            filtered = filtered[filtered.get("priority","") == prio_filter]
        if products_filter != "Tous Produits":
            filtered = filtered[filtered.get("product","") == products_filter]

    st.markdown("<div style='height:16px;'></div>", unsafe_allow_html=True)

    # Table
    st.markdown(f"""
        <div class="table-wrap">
            <div class="table-top">
                <span style="font-family:'Syne',sans-serif;font-size:14px;font-weight:700;color:#1A1F1C;">Base de données</span>
                <span class="table-count">{len(filtered)} résultat{'s' if len(filtered)!=1 else ''}</span>
            </div>
            <div class="th-row">
                <span class="th-cell">Société</span>
                <span class="th-cell">Priorité</span>
                <span class="th-cell">Produit</span>
                <span class="th-cell">Analyse</span>
                <span class="th-cell">Decision Maker</span>
                <span class="th-cell">Email</span>
                <span class="th-cell">Website</span>
                <span class="th-cell"></span>
            </div>
    """, unsafe_allow_html=True)

    if filtered.empty:
        st.markdown("""
            <div class="empty-state">
                <div class="empty-icon">🌿</div>
                <div class="empty-title">Aucun prospect</div>
                <div class="empty-sub">Ajoutez votre premier prospect avec le bouton "Nouveau Prospect"</div>
            </div>
        """, unsafe_allow_html=True)
    else:
        for _, row in filtered.iterrows():
            cols = st.columns([2, 0.7, 1, 1.4, 1.2, 1.3, 1, 0.6])

            with cols[0]:
                st.markdown(f'<span class="td-company">{row.get("company","—")}</span>', unsafe_allow_html=True)
            with cols[1]:
                st.markdown(priority_badge(row.get("priority","Low")), unsafe_allow_html=True)
            with cols[2]:
                st.markdown(f'<span class="td-text">{row.get("product","—")}</span>', unsafe_allow_html=True)
            with cols[3]:
                analyse_text = row.get("analyse","—") or "—"
                st.markdown(f'<span class="td-analyse">{analyse_text}</span>', unsafe_allow_html=True)
            with cols[4]:
                st.markdown(f'<span class="td-text">{row.get("decision_maker","—")}</span>', unsafe_allow_html=True)
            with cols[5]:
                email_val = row.get("email","") or ""
                if email_val:
                    st.markdown(f'<a href="mailto:{email_val}" class="td-link">{email_val}</a>', unsafe_allow_html=True)
                else:
                    st.markdown('<span class="td-text">—</span>', unsafe_allow_html=True)
            with cols[6]:
                website_val = row.get("website","") or ""
                if website_val:
                    url_display = website_val.replace("https://","").replace("http://","").rstrip("/")
                    st.markdown(f'<a href="{website_val}" target="_blank" class="td-link">↗ {url_display[:22]}</a>', unsafe_allow_html=True)
                else:
                    st.markdown('<span class="td-text">—</span>', unsafe_allow_html=True)
            with cols[7]:
                if st.button("✎", key=f"edit_{row['id']}", help="Modifier", use_container_width=True):
                    st.session_state["edit_pid"]  = row["id"]
                    st.session_state["edit_data"] = row.to_dict()
                    st.rerun()

    st.markdown("</div>", unsafe_allow_html=True)

    # Open "new" modal
    if st.session_state.pop("open_new", False):
        prospect_modal()

    # Open "edit" modal
    if "edit_pid" in st.session_state:
        prospect_modal(pid=st.session_state["edit_pid"], initial_data=st.session_state.get("edit_data",{}))

# =============================================================================
# PAGE: VEILLE IA (GEMINI)
# =============================================================================

def page_veille():
    st.markdown("""
        <div class="page-header">
            <h1 class="page-title">Veille Marché IA</h1>
            <p class="page-subtitle">Monitoring sulforaphane · brocoli · BAD · compléments alimentaires</p>
        </div>
    """, unsafe_allow_html=True)

    tab1, tab2 = st.tabs(["🔬 Analyse marché", "📰 Actualités & Tendances"])

    with tab1:
        st.markdown("""
            <div style="background:#F0FDF4;border:1px solid #BBF7D0;border-radius:12px;padding:18px 22px;margin-bottom:24px;">
                <p style="font-size:13px;color:#166534;margin:0;line-height:1.6;">
                    <strong>Sulfodyne (L-sulforaphane stable · Ingood by Olga)</strong><br>
                    Posez une question sur le marché, les concurrents, les opportunités ou les tendances BAD/sulforaphane.
                </p>
            </div>
        """, unsafe_allow_html=True)

        prompt_suggestions = [
            "Quelles sont les tendances 2024-2025 des BADs à base de sulforaphane en Europe ?",
            "Quels sont les principaux concurrents sur le marché des ingrédients sulforaphane B2B ?",
            "Quels segments (sport, senior, femme, beauté) sont les plus porteurs pour les BADs brocoli ?",
            "Quelle est la réglementation EFSA pour les allégations santé sulforaphane ?",
        ]

        st.markdown('<span class="form-label">QUESTION / REQUÊTE</span>', unsafe_allow_html=True)
        
        col_sug, _ = st.columns([3,1])
        with col_sug:
            selected_sug = st.selectbox("sug", ["Choisir une suggestion…"] + prompt_suggestions, label_visibility="collapsed", key="sug_box")

        user_prompt = st.text_area(
            "prompt",
            value=selected_sug if selected_sug != "Choisir une suggestion…" else "",
            height=90,
            label_visibility="collapsed",
            placeholder="Posez votre question sur le marché sulforaphane / BAD…",
            key="veille_prompt"
        )

        if st.button("🔬 Analyser avec Gemini", type="primary"):
            if not user_prompt.strip():
                st.warning("Entrez une question")
            else:
                with st.spinner("Analyse en cours…"):
                    try:
                        model = genai.GenerativeModel("gemini-1.5-flash")
                        system_ctx = """Tu es un expert en Business Development dans le secteur des ingrédients nutraceutiques B2B.
                        Tu travailles pour Ingood by Olga, société française spécialisée dans le Sulfodyne (L-sulforaphane stable issu du brocoli).
                        Réponds en français, de manière professionnelle, avec des données concrètes et actionnables pour une BizDev.
                        Focus sur : marché BAD (Biologically Active Dietary supplements), compléments alimentaires, nutraceutique, santé fonctionnelle.
                        Mentionne des chiffres de marché, des tendances, des opportunités de prospection quand c'est pertinent."""
                        
                        response = model.generate_content(f"{system_ctx}\n\nQuestion : {user_prompt}")
                        
                        st.markdown("""
                            <div style="background:white;border:1px solid #E8EAE6;border-radius:12px;padding:24px;margin-top:20px;">
                                <div style="font-family:'Syne',sans-serif;font-size:13px;font-weight:700;color:#1E3F35;margin-bottom:14px;text-transform:uppercase;letter-spacing:0.5px;">
                                    🤖 Analyse Gemini
                                </div>
                        """, unsafe_allow_html=True)
                        st.markdown(response.text)
                        st.markdown("</div>", unsafe_allow_html=True)
                    except Exception as e:
                        st.error(f"Erreur Gemini : {e}")

    with tab2:
        st.markdown("""
            <div style="margin-bottom:20px;">
                <p style="font-size:13px;color:#6B7280;">Générez un brief de veille automatique sur les dernières actualités sulforaphane, brocoli et BAD.</p>
            </div>
        """, unsafe_allow_html=True)

        c1, c2, c3 = st.columns(3)
        with c1:
            topic_focus = st.selectbox("focus", [
                "Marché BAD sulforaphane global",
                "Nouveaux lancements produits brocoli/sulforaphane",
                "Études cliniques sulforaphane récentes",
                "Réglementation compléments alimentaires EU",
                "Opportunités marché Asie (Russie, Chine)",
                "Concurrents ingrédients brocoli B2B",
            ], label_visibility="collapsed")
        with c2:
            geo_focus = st.selectbox("geo", ["Europe", "Monde entier", "Asie", "Amérique du Nord", "Russie & CEI"], label_visibility="collapsed")
        with c3:
            if st.button("📰 Générer la veille", type="primary", use_container_width=True):
                with st.spinner("Génération du brief de veille…"):
                    try:
                        model = genai.GenerativeModel("gemini-1.5-flash")
                        brief_prompt = f"""En tant qu'expert nutraceutique B2B, génère un brief de veille marché structuré sur : "{topic_focus}" pour la zone géographique : {geo_focus}.
                        
                        Contexte : Ingood by Olga commercialise Sulfodyne, un L-sulforaphane stable issu de graines de brocoli, ingrédient premium B2B pour BADs et compléments alimentaires.
                        
                        Format souhaité :
                        ## 📊 Tendances clés
                        ## 🌟 Opportunités de marché
                        ## ⚠️ Points de vigilance
                        ## 🎯 Recommandations BizDev
                        
                        Sois précis, factuel, orienté action commerciale B2B. Réponds en français."""
                        
                        response = model.generate_content(brief_prompt)
                        st.session_state["veille_brief"] = response.text
                    except Exception as e:
                        st.error(f"Erreur : {e}")

        if "veille_brief" in st.session_state:
            st.markdown("""
                <div style="background:white;border:1px solid #E8EAE6;border-radius:12px;padding:24px;margin-top:20px;">
                    <div style="font-family:'Syne',sans-serif;font-size:13px;font-weight:700;color:#1E3F35;margin-bottom:14px;text-transform:uppercase;letter-spacing:0.5px;">
                        📰 Brief de Veille
                    </div>
            """, unsafe_allow_html=True)
            st.markdown(st.session_state["veille_brief"])
            st.markdown("</div>", unsafe_allow_html=True)

# =============================================================================
# PAGE: IMPORT / EXPORT
# =============================================================================

def page_excel():
    st.markdown("""
        <div class="page-header">
            <h1 class="page-title">Import · Export</h1>
            <p class="page-subtitle">Gérez votre base de données prospects</p>
        </div>
    """, unsafe_allow_html=True)

    df = load_prospects()

    col_ex, col_im = st.columns(2, gap="large")

    # ── EXPORT ──
    with col_ex:
        with st.container(border=True):
            st.markdown("""
                <p style="font-family:'Syne',sans-serif;font-size:16px;font-weight:700;color:#1A1F1C;margin:0 0 6px;">
                    📤  Export Excel
                </p>
                <p style="font-size:13px;color:#8A9490;margin:0 0 20px;">
                    Téléchargez toute votre base de données en format .xlsx
                </p>
            """, unsafe_allow_html=True)

            if df.empty:
                st.info("Aucun prospect à exporter")
            else:
                # Clean export columns
                export_cols = ["company","priority","product","analyse","decision_maker","email","website","created_at","updated_at"]
                export_df = df[[c for c in export_cols if c in df.columns]].copy()
                export_df.columns = [c.replace("_"," ").title() for c in export_df.columns]

                buffer = BytesIO()
                with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                    export_df.to_excel(writer, index=False, sheet_name="Prospects Sulfodyne")

                    # Style header
                    wb = writer.book
                    ws = writer.sheets["Prospects Sulfodyne"]
                    from openpyxl.styles import PatternFill, Font
                    green_fill = PatternFill("solid", fgColor="1E3F35")
                    for cell in ws[1]:
                        cell.fill = green_fill
                        cell.font = Font(color="FFFFFF", bold=True)
                    for col in ws.columns:
                        ws.column_dimensions[col[0].column_letter].width = 22

                buffer.seek(0)
                today = datetime.now().strftime("%Y%m%d")
                st.download_button(
                    f"⬇  Télécharger Excel ({len(df)} lignes)",
                    data=buffer,
                    file_name=f"sulfodyne_prospects_{today}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True
                )

    # ── IMPORT ──
    with col_im:
        with st.container(border=True):
            st.markdown("""
                <p style="font-family:'Syne',sans-serif;font-size:16px;font-weight:700;color:#1A1F1C;margin:0 0 6px;">
                    📥  Import Excel
                </p>
                <p style="font-size:13px;color:#8A9490;margin:0 0 20px;">
                    Importez un fichier .xlsx. Les colonnes attendues :<br>
                    <code style="font-size:11px;">company, priority, product, analyse, decision_maker, email, website</code>
                </p>
            """, unsafe_allow_html=True)

            uploaded = st.file_uploader("Fichier Excel", type=["xlsx"], label_visibility="collapsed")
            if uploaded:
                try:
                    import_df = pd.read_excel(uploaded)
                    import_df.columns = [c.strip().lower().replace(" ","_") for c in import_df.columns]

                    st.markdown(f"**{len(import_df)} lignes** détectées :")
                    st.dataframe(import_df.head(5), use_container_width=True)

                    if st.button("✓ Importer dans la base", type="primary", use_container_width=True, key="do_import"):
                        ok = err = 0
                        for _, r in import_df.iterrows():
                            try:
                                payload = {
                                    "company":        str(r.get("company","")).upper().strip(),
                                    "priority":       str(r.get("priority","Low")).strip(),
                                    "product":        str(r.get("product","")).strip(),
                                    "analyse":        str(r.get("analyse","")).strip(),
                                    "decision_maker": str(r.get("decision_maker","")).strip(),
                                    "email":          str(r.get("email","")).strip(),
                                    "website":        str(r.get("website","")).strip(),
                                    "created_at":     datetime.now().isoformat(),
                                    "updated_at":     datetime.now().isoformat(),
                                }
                                if payload["priority"] not in PRIORITY_OPTS:
                                    payload["priority"] = "Low"
                                if payload["company"]:
                                    db().table("sulfodyne_prospects").insert(payload).execute()
                                    ok += 1
                            except:
                                err += 1
                        refresh()
                        st.success(f"✓ {ok} prospects importés · {err} erreurs")
                        st.rerun()
                except Exception as e:
                    st.error(f"Erreur lecture fichier : {e}")

    # Preview table
    st.markdown("<div style='height:32px;'></div>", unsafe_allow_html=True)
    if not df.empty:
        st.markdown("""
            <p style="font-family:'Syne',sans-serif;font-size:15px;font-weight:700;color:#1A1F1C;margin-bottom:12px;">
                Aperçu de la base complète
            </p>
        """, unsafe_allow_html=True)
        preview_cols = ["company","priority","product","decision_maker","email","website"]
        st.dataframe(
            df[[c for c in preview_cols if c in df.columns]].rename(
                columns={"company":"Société","priority":"Priorité","product":"Produit",
                         "decision_maker":"Decision Maker","email":"Email","website":"Website"}
            ),
            use_container_width=True,
            height=320
        )

# =============================================================================
# MAIN
# =============================================================================

def main():
    # Auth simple
    if not st.session_state.get("auth"):
        pwd_ok = st.secrets.get("ACCESS_PASSWORD","")
        token_ok = st.secrets.get("ACCESS_TOKEN","")

        if st.query_params.get("token") == token_ok and token_ok:
            st.session_state["auth"] = True
        else:
            _, col, _ = st.columns([1,1,1])
            with col:
                st.markdown("<div style='height:80px;'></div>", unsafe_allow_html=True)
                st.markdown("""
                    <div style="text-align:center;padding:40px 32px;background:white;border-radius:16px;border:1px solid #E8EAE6;box-shadow:0 8px 32px rgba(0,0,0,0.08);">
                        <div style="font-family:'Syne',sans-serif;font-size:22px;font-weight:800;color:#1E3F35;margin-bottom:4px;">Ingood by Olga</div>
                        <div style="font-size:13px;color:#8A9490;margin-bottom:28px;">Sulfodyne · Business Development</div>
                    </div>
                """, unsafe_allow_html=True)
                pwd = st.text_input("", type="password", placeholder="Mot de passe", label_visibility="collapsed")
                if st.button("Accéder →", type="primary", use_container_width=True):
                    if pwd == pwd_ok:
                        st.session_state["auth"] = True
                        st.rerun()
                    else:
                        st.error("Mot de passe incorrect")
            return

    if not db():
        st.error("Connexion base de données impossible. Vérifiez vos secrets Supabase.")
        st.stop()

    page = render_sidebar()

    if page == "Prospects":
        page_prospects()
    elif page == "Veille IA":
        page_veille()
    elif page == "Import/Export":
        page_excel()

if __name__ == "__main__":
    main()
